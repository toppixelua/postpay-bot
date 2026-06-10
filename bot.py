import os, json, base64, re, io
from datetime import datetime
import httpx
import openpyxl
import psycopg2
from psycopg2.extras import Json
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

BOT_TOKEN     = os.environ["BOT_TOKEN"]
ANTHROPIC_KEY = os.environ["ANTHROPIC_KEY"]
DATABASE_URL  = os.environ.get("DATABASE_URL", "")

# ── Назви типів відомостей ────────────────────────────────────────────────────
VED_NAMES = {
    "01": "Пенсія",
    "02": "Доплата до пенсії",
    "13": "Одноразова допомога дітей",
    "14": "Соціальна допомога до пенсії",
    "28": "Допомога по догляду",
    "30": "Компенсація дітям від ЧАЕС",
    "32": "Компенсація за харчування",
    "52": "Багатодітним",
    "57": "Малозабезпечені сім'ї",
    "58": "З інвалідністю дитинства",
    "98": "Житлова субсидія",
    "99": "Пільги",
}

def ved_name(type_str):
    t = re.sub(r'[^0-9]', '', str(type_str or '')).zfill(2)
    return VED_NAMES.get(t, "В" + t)

# ── Database ──────────────────────────────────────────────────────────────────
def get_conn():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS vedomosti (
                    id SERIAL PRIMARY KEY,
                    user_id BIGINT NOT NULL,
                    vedomist_type TEXT,
                    dilinitsa TEXT,
                    period TEXT,
                    source_file TEXT,
                    rows JSONB,
                    created_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(user_id, vedomist_type, dilinitsa, period)
                );
                CREATE TABLE IF NOT EXISTS phone_books (
                    user_id BIGINT PRIMARY KEY,
                    data JSONB,
                    updated_at TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS paid_marks (
                    user_id BIGINT NOT NULL,
                    person_key TEXT NOT NULL,
                    pay_date TEXT NOT NULL,
                    paid_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (user_id, person_key, pay_date)
                );
                CREATE TABLE IF NOT EXISTS account_links (
                    user_id BIGINT NOT NULL,
                    account TEXT NOT NULL,
                    pb_entry_id INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (user_id, account)
                );
            """)
            try:
                cur.execute("ALTER TABLE paid_marks ADD COLUMN IF NOT EXISTS pay_date TEXT NOT NULL DEFAULT ''")
            except Exception:
                pass
        conn.commit()

def db_get_vedomosti(uid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT vedomist_type, dilinitsa, period, source_file, rows FROM vedomosti WHERE user_id=%s ORDER BY created_at", (uid,))
            rows = cur.fetchall()
    return [{"vedomist_type":r[0],"dilinitsa":r[1],"period":r[2],"source_file":r[3],"rows":r[4]} for r in rows]

def db_save_vedomist(uid, v):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO vedomosti (user_id, vedomist_type, dilinitsa, period, source_file, rows)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT (user_id, vedomist_type, dilinitsa, period)
                DO UPDATE SET rows=EXCLUDED.rows, source_file=EXCLUDED.source_file, created_at=NOW()
            """, (uid, v.get("vedomist_type"), v.get("dilinitsa"), v.get("period"), v.get("source_file"), Json(v.get("rows",[]))))
        conn.commit()

def db_clear_vedomosti(uid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM vedomosti WHERE user_id=%s", (uid,))
        conn.commit()

def db_get_phones(uid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM phone_books WHERE user_id=%s", (uid,))
            row = cur.fetchone()
    return row[0] if row else {}

def db_save_phones(uid, data):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO phone_books (user_id, data, updated_at)
                VALUES (%s,%s,NOW())
                ON CONFLICT (user_id) DO UPDATE SET data=EXCLUDED.data, updated_at=NOW()
            """, (uid, Json(data)))
        conn.commit()

def db_get_paid(uid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT person_key, pay_date FROM paid_marks WHERE user_id=%s", (uid,))
            return {(row[0], row[1]) for row in cur.fetchall()}

def db_mark_paid(uid, key, pay_date):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO paid_marks (user_id, person_key, pay_date) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                (uid, key, pay_date)
            )
        conn.commit()

def db_unmark_paid(uid, key, pay_date=None):
    with get_conn() as conn:
        with conn.cursor() as cur:
            if pay_date:
                cur.execute("DELETE FROM paid_marks WHERE user_id=%s AND person_key=%s AND pay_date=%s", (uid, key, pay_date))
            else:
                cur.execute("DELETE FROM paid_marks WHERE user_id=%s AND person_key=%s", (uid, key))
        conn.commit()

def db_clear_paid(uid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM paid_marks WHERE user_id=%s", (uid,))
        conn.commit()

# ── Account links ─────────────────────────────────────────────────────────────
def db_link_account(uid, account, pb_entry_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO account_links (user_id, account, pb_entry_id)
                VALUES (%s,%s,%s)
                ON CONFLICT (user_id, account) DO UPDATE SET pb_entry_id=EXCLUDED.pb_entry_id, created_at=NOW()
            """, (uid, account, pb_entry_id))
        conn.commit()

def db_unlink_account(uid, account):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM account_links WHERE user_id=%s AND account=%s", (uid, account))
        conn.commit()

def db_get_links(uid):
    """Повертає dict {account: pb_entry_id}"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT account, pb_entry_id FROM account_links WHERE user_id=%s", (uid,))
            return {row[0]: row[1] for row in cur.fetchall()}

def db_get_pb_entry(uid, entry_id):
    """Отримати один запис pb_entries по id"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, street, name, phone, notes FROM pb_entries WHERE id=%s AND user_id=%s",
                (entry_id, uid)
            )
            return cur.fetchone()

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm_addr(addr):
    return re.sub(r'[^А-ЯІЇЄҐа-яіїєґA-Za-z0-9]', '',
        addr.upper()
        .replace('ВУЛ.','').replace('ВУЛИЦЯ','').replace('СОШИЧНЕ','')
        .replace(';;','').replace('Б/Н','').replace('М.КАМІНЬ-КАШИРСЬКИЙ','')
        .replace('С.ЩИТИНЬ','').replace('С.СОШИЧНЕ',''))

def make_person_key(r):
    acc = str(r.get("account","") or "").strip()
    if acc:
        return "ACC|" + acc
    return r["name"] + "|" + norm_addr(r.get("address",""))

def norm_date(s):
    m = re.match(r'(\d{1,2})[.\-/](\d{1,2})', str(s))
    if not m: return str(s)
    return str(int(m.group(1))).zfill(2) + "." + str(int(m.group(2))).zfill(2)

def parse_date(s):
    m = re.match(r'(\d{1,2})[.\-/](\d{1,2})', str(s))
    if not m: return None
    try: return datetime(datetime.now().year, int(m.group(2)), int(m.group(1)))
    except: return None

def today_str():
    return datetime.now().strftime("%d.%m")

def is_payable_by(pay_date, limit_date):
    d = parse_date(pay_date)
    if not d: return False
    return d <= limit_date

def fmt_hrn(n):
    try: return "{:,.2f} грн".format(float(n)).replace(',', ' ')
    except: return str(n)

def parse_xlsx(data):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    phone_map = {}
    for row in ws.iter_rows(values_only=True):
        col1 = str(row[1] if len(row) > 1 else '').strip()
        col3 = str(row[3] if len(row) > 3 else '').strip()
        if 'вул.' in col1.lower() and col3 and col3 != 'None':
            key = norm_addr(col1)
            if key and key not in phone_map:
                phone_map[key] = col3
    return phone_map

def extract_json(text):
    text = re.sub(r'^```(?:json)?\s*', '', text.strip())
    text = re.sub(r'\s*```$', '', text.strip()).strip()
    try: return json.loads(text)
    except: pass
    s, e = text.find('{'), text.rfind('}')
    if s != -1 and e > s:
        try: return json.loads(text[s:e+1])
        except: pass
    return None

# ── find_phone: account_links → pb_entries → старий dict ──
def find_phone(phone_book, addr, uid=None, account=None, links=None):
    """Шукає телефон: 1) по прив'язці account→pb_entry, 2) по адресі в pb_entries, 3) старий dict"""
    # 1) Пряма прив'язка account → pb_entry_id
    if uid is not None and account and links:
        entry_id = links.get(str(account).strip())
        if entry_id:
            try:
                entry = db_get_pb_entry(uid, entry_id)
                if entry and entry[3] and entry[3].strip():
                    return entry[3].strip()
            except Exception:
                pass

    # 2) Пошук по адресі в pb_entries
    if uid is not None:
        try:
            entries = pb_find_by_street(uid, addr)
            for eid, street, name, phone, notes in entries:
                if phone and phone.strip():
                    return phone.strip()
        except Exception:
            pass

    # 3) Fallback на старий phone_book dict
    if not phone_book:
        return ''
    key = norm_addr(addr)
    if not key:
        return ''
    if key in phone_book:
        return phone_book[key]
    for k, v in phone_book.items():
        if len(key) > 5 and (k.startswith(key) or key.startswith(k)):
            return v
    return ''

def build_all_rows(vedomosti, phone_book, uid=None):
    # Завантажити прив'язки один раз
    links = db_get_links(uid) if uid is not None else {}
    grouped = {}
    for v in vedomosti:
        for r in v.get("rows", []):
            key = make_person_key(r)
            acc = str(r.get("account","") or "").strip()
            if key not in grouped:
                grouped[key] = {
                    "name": r["name"],
                    "address": r.get("address",""),
                    "passport": r.get("passport",""),
                    "account": acc,
                    "phone": find_phone(phone_book, r.get("address",""), uid=uid, account=acc, links=links),
                    "linked": acc in links,
                    "veds": []
                }
            grouped[key]["veds"].append({
                "type": v.get("vedomist_type","?"),
                "dil":  v.get("dilinitsa","?"),
                "sum":  r.get("sum", 0),
                "pay_date": norm_date(r.get("pay_date","?"))
            })
    return list(grouped.values())

def is_person_fully_paid(r, paid_set):
    key = make_person_key(r)
    return all((key, v["pay_date"]) in paid_set for v in r["veds"])

def is_ved_paid(r, ved, paid_set):
    key = make_person_key(r)
    return (key, ved["pay_date"]) in paid_set

def format_rows(rows, phone_book, paid_set=None, uid=None):
    if paid_set is None:
        paid_set = set()
    chunks = []
    current = ""
    for i, r in enumerate(rows, 1):
        key       = make_person_key(r)
        addr      = r["address"].replace("Сошичне, ","").replace(";;","").strip()
        phone     = r.get("phone","") or find_phone(phone_book, r["address"], uid=uid)
        account   = r.get("account","")
        all_paid  = is_person_fully_paid(r, paid_set)

        linked = r.get("linked", False)
        if all_paid:
            line = str(i) + ". ✅ " + r["name"] + "\n"
        else:
            line = str(i) + ". " + r["name"] + "\n"
        line += "   📍 " + addr + "\n"
        if phone:
            link_icon = "🔗" if linked else "📞"
            line += "   " + link_icon + " " + phone + "\n"
        if account:
            line += "   🔢 " + account + "\n"

        for ved in r["veds"]:
            ved_paid = is_ved_paid(r, ved, paid_set)
            status = "✅" if ved_paid else "⏳"
            line += "   " + status + " В" + ved["type"] + "·" + ved["dil"]
            line += " · " + fmt_hrn(ved.get("sum",0))
            line += " · 📅 " + ved["pay_date"] + "\n"

        total_sum  = sum(float(v.get("sum",0)) for v in r["veds"])
        paid_sum   = sum(float(v.get("sum",0)) for v in r["veds"] if is_ved_paid(r, v, paid_set))
        unpaid_sum = total_sum - paid_sum

        if all_paid:
            line += "   ✔️ Виплачено: " + fmt_hrn(total_sum) + "\n"
        else:
            line += "   💳 До виплати: " + fmt_hrn(unpaid_sum) + "\n"
            if paid_sum > 0:
                line += "   ✅ Виплачено: " + fmt_hrn(paid_sum) + "\n"

        line += "   🪪 " + r["passport"] + "\n\n"

        if len(current) + len(line) > 3800:
            chunks.append(current)
            current = line
        else:
            current += line
    if current:
        chunks.append(current)
    return chunks

def find_by_account_or_name(all_rows, query):
    q = query.strip()
    by_acc = [r for r in all_rows if r.get("account","") == q]
    if by_acc: return by_acc
    return [r for r in all_rows if q.upper() in r["name"].upper()]

# ── Claude API ────────────────────────────────────────────────────────────────
async def ask_claude(b64):
    prompt = (
        "Це відомість ПФУ на виплату пенсій. Зчитай всі рядки таблиці з УСІХ сторінок.\n"
        "Поверни ТІЛЬКИ валідний JSON без markdown, без пояснень:\n"
        '{"vedomist_type":"99","period":"Червень 2026 01","dilinitsa":"50","rows":[{"account":"001256811781","name":"БАЛЕЦЬКИЙ ВОЛОДИМИР ПОРФИРІЙОВИЧ","address":"Сошичне, МИРУ, 94;;","sum":2301.28,"passport":"AC 000393421","pay_date":"04.06"}]}\n'
        "- account: номер особового рахунку з першої колонки\n"
        "- vedomist_type: число з рядка Тип виплати\n"
        "- dilinitsa: перший номер після назви села\n"
        "- period: текст після слова період\n"
        "- name: ПІБ ВЕЛИКИМИ ЛІТЕРАМИ\n"
        "- address: адреса як є\n"
        "- sum: число\n"
        "- pay_date: DD.MM"
    )
    async with httpx.AsyncClient(timeout=180) as client:
        resp = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "anthropic-beta": "pdfs-2024-09-25",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 8192,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                        {"type": "text", "text": prompt}
                    ]
                }]
            }
        )
        resp.raise_for_status()
        data = resp.json()
        text = "".join(c.get("text","") for c in data.get("content",[]))
        parsed = extract_json(text)
        if not parsed:
            raise ValueError("JSON не знайдено. Відповідь: " + text[:300])
        if not isinstance(parsed.get("rows"), list):
            parsed["rows"] = []
        return parsed

# ── Handlers ──────────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привіт! Бот для обробки відомостей ПФУ.\n\n"
        "📄 Надішли PDF — додається до загального списку\n"
        "📞 Надішли XLSX — завантажує телефонну книгу\n\n"
        "Команди:\n"
        "/status — статус відомостей\n"
        "/list — зведений список\n"
        "/today 07.06 — виплати до дати\n"
        "/search Шв — пошук (2+ літери)\n"
        "/multi — отримувачі з 2+ виплатами\n"
        "/clear — очистити відомості\n"
        "/phones — телефонна книга\n\n"
        "Відмітки виплат:\n"
        "/paid 001256811781 — всі виплати що настали\n"
        "/paid 001256811781 05.06 — конкретна дата\n"
        "/unpaid 001256811781 — зняти всі позначки\n"
        "/unpaid 001256811781 05.06 — зняти по даті\n"
        "/unpaid_list — невиплачені\n"
        "/clear_paid — зняти всі позначки\n\n"
        "Телефонна книга:\n"
        "/pb_find Шкільна 7 — пошук по адресі\n"
        "/pb_find Левчук — пошук по ПІБ\n"
        "/pb_phone Зарічна 36 0991234567 — швидко додати телефон\n"
        "/pb_add вул. Адреса | ПІБ | телефон — повний запис\n"
        "/pb_edit [ID] phone 0991234567 — редагувати\n"
        "/pb_del [ID] — видалити\n"
        "/pb_export — скачати xlsx"
    )

async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає. Надішли PDF.")
        return
    rows = build_all_rows(veds, phones, uid=uid)
    paid_set = db_get_paid(uid)
    total_veds_count = sum(len(r["veds"]) for r in rows)
    paid_count = sum(1 for r in rows for v in r["veds"] if is_ved_paid(r, v, paid_set))
    total_sum = sum(sum(float(v.get("sum",0)) for v in r["veds"]) for r in rows)
    ved_list = "\n".join([
        "  • В" + v.get("vedomist_type","?") + " " + ved_name(v.get("vedomist_type","")) +
        " ·Діл." + v.get("dilinitsa","?") +
        " — " + str(len(v.get("rows",[]))) + " ос. (" + v.get("period","") + ")"
        "  →  /list " + re.sub(r'[^0-9]','',v.get("vedomist_type","?"))
        for v in veds
    ])
    await update.message.reply_text(
        "📊 Статус:\n" + ved_list + "\n\n"
        "👥 Унікальних осіб: " + str(len(rows)) + "\n"
        "✅ Виплат здійснено: " + str(paid_count) + "/" + str(total_veds_count) + "\n"
        "💰 Загальна сума: " + fmt_hrn(total_sum) + "\n"
        "📞 Телефонна книга: " + str(len(phones)) + " адрес\n\n"
        "Фільтр: /list · /list В01 · /list 50 · /list В01 50"
    )

async def cmd_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає. Надішли PDF.")
        return

    # Парсимо аргументи фільтру: /list [тип] [дільниця]
    # Приклади: /list В01  /list 50  /list В01 50  /list 1 50
    filter_type = None
    filter_dil  = None
    if ctx.args:
        for arg in ctx.args:
            a = arg.upper().lstrip('ВB')  # strip В/B prefix
            # Якщо це число — може бути тип або дільниця
            # Визначаємо: якщо вже є filter_type — це дільниця, інакше тип
            raw = re.sub(r'[^0-9]', '', arg)
            if arg.upper().startswith(('В', 'B')):
                filter_type = raw
            elif raw:
                if filter_type is None:
                    filter_type = raw
                else:
                    filter_dil = raw
            else:
                # не число — спробувати як дільницю (Діл.50 → 50)
                m = re.search(r'\d+', arg)
                if m:
                    filter_dil = m.group()

    # Якщо є фільтр — відфільтрувати відомості
    if filter_type or filter_dil:
        filtered_veds = []
        for v in veds:
            match_type = (not filter_type) or (re.sub(r'[^0-9]','', v.get("vedomist_type","")) == filter_type)
            match_dil  = (not filter_dil)  or (re.sub(r'[^0-9]','', v.get("dilinitsa",""))  == filter_dil)
            if match_type and match_dil:
                filtered_veds.append(v)
        if not filtered_veds:
            # Показати доступні відомості
            ved_list = "\n".join([
                "  /list " + re.sub(r'[^0-9]','',v.get("vedomist_type","?")) +
                " — В" + v.get("vedomist_type","?") + " " + ved_name(v.get("vedomist_type","")) +
                " ·Діл." + v.get("dilinitsa","?") +
                " · " + str(len(v.get("rows",[]))) + " ос."
                for v in veds
            ])
            await update.message.reply_text(
                "🔍 Нічого не знайдено за фільтром.\n\n"
                "Доступні відомості:\n" + ved_list +
                "\n\nПриклади: /list 52 · /list 99 · /list 01"
            )
            return
        if filter_type:
            label = "В" + filter_type + " " + ved_name(filter_type)
        else:
            label = ""
        if filter_dil: label += ("·" if label else "") + "Діл." + filter_dil
        veds_to_show = filtered_veds
    else:
        label = "зведений"
        veds_to_show = veds

    rows = build_all_rows(veds_to_show, phones, uid=uid)
    rows.sort(key=lambda r: r["veds"][0].get("pay_date","99.99"))
    paid_set = db_get_paid(uid)
    unpaid = sum(1 for r in rows if not is_person_fully_paid(r, paid_set))
    header = "📋 Список " + label + " · " + str(len(rows)) + " осіб · не виплачено: " + str(unpaid) + "\n" + "─"*28
    await update.message.reply_text(header)
    for chunk in format_rows(rows, phones, paid_set, uid=uid):
        await update.message.reply_text(chunk)

async def cmd_today(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає. Надішли PDF.")
        return
    if not ctx.args:
        await update.message.reply_text("⚠️ Вкажи дату: /today 07.06")
        return
    limit = parse_date(ctx.args[0])
    if not limit:
        await update.message.reply_text("⚠️ Невірний формат. Приклад: /today 07.06")
        return
    all_rows = build_all_rows(veds, phones, uid=uid)
    filtered = []
    for r in all_rows:
        payable = [v for v in r["veds"] if is_payable_by(v["pay_date"], limit)]
        if payable:
            filtered.append(dict(r, veds=payable))
    if not filtered:
        await update.message.reply_text("📭 Немає виплат до " + ctx.args[0])
        return
    filtered.sort(key=lambda r: r["veds"][0].get("pay_date","99.99"))
    paid_set = db_get_paid(uid)
    total = sum(sum(float(v.get("sum",0)) for v in r["veds"]) for r in filtered)
    unpaid = sum(1 for r in filtered if not is_person_fully_paid(r, paid_set))
    await update.message.reply_text("📅 Виплати до " + ctx.args[0] + " · " + str(len(filtered)) + " осіб · не виплачено: " + str(unpaid) + "\n💰 " + fmt_hrn(total) + "\n" + "─"*28)
    for chunk in format_rows(filtered, phones, paid_set, uid=uid):
        await update.message.reply_text(chunk)

async def cmd_search(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає. Надішли PDF.")
        return
    if not ctx.args:
        await update.message.reply_text("⚠️ Вкажи частину прізвища: /search Шв")
        return
    query = " ".join(ctx.args).upper().strip()
    if len(query) < 2:
        await update.message.reply_text("⚠️ Мінімум 2 літери")
        return
    found = [r for r in build_all_rows(veds, phones, uid=uid) if query in r["name"].upper()]
    if not found:
        await update.message.reply_text("🔍 Нічого не знайдено: " + query)
        return
    paid_set = db_get_paid(uid)
    await update.message.reply_text("🔍 '" + query + "' · " + str(len(found)) + " осіб\n" + "─"*28)
    for chunk in format_rows(found, phones, paid_set, uid=uid):
        await update.message.reply_text(chunk)

async def cmd_multi(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає. Надішли PDF.")
        return
    all_rows = build_all_rows(veds, phones, uid=uid)
    multi = [r for r in all_rows if len(r["veds"]) >= 2]
    if not multi:
        await update.message.reply_text("📭 Немає отримувачів з двома і більше виплатами.")
        return
    multi.sort(key=lambda r: r["name"])
    paid_set = db_get_paid(uid)
    total = sum(sum(float(v.get("sum",0)) for v in r["veds"]) for r in multi)
    await update.message.reply_text("👥 Отримувачів з 2+ виплатами: " + str(len(multi)) + "\n💰 Загальна сума: " + fmt_hrn(total) + "\n" + "─"*28)
    for chunk in format_rows(multi, phones, paid_set, uid=uid):
        await update.message.reply_text(chunk)

async def cmd_paid(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає.")
        return
    if not ctx.args:
        await update.message.reply_text(
            "⚠️ Вкажи номер рахунку або прізвище:\n"
            "/paid 001256811781 — всі виплати що настали\n"
            "/paid 001256811781 05.06 — конкретна дата"
        )
        return
    query = ctx.args[0].strip()
    specific_date = norm_date(ctx.args[1]) if len(ctx.args) >= 2 else None
    all_rows = build_all_rows(veds, phones, uid=uid)
    found = find_by_account_or_name(all_rows, query)
    if not found:
        await update.message.reply_text("🔍 Не знайдено: " + query)
        return
    if len(found) > 1:
        names = "\n".join([str(i+1) + ". " + r.get("account","") + " " + r["name"] for i, r in enumerate(found)])
        await update.message.reply_text("Знайдено кілька — вкажи номер рахунку:\n" + names)
        return
    r = found[0]
    key = make_person_key(r)
    today = today_str()
    marked = []
    if specific_date:
        matching = [v for v in r["veds"] if norm_date(v["pay_date"]) == specific_date]
        if not matching:
            await update.message.reply_text("⚠️ Виплати на " + specific_date + " не знайдено для " + r["name"])
            return
        db_mark_paid(uid, key, specific_date)
        marked = [specific_date]
    else:
        limit = parse_date(today)
        for v in r["veds"]:
            if is_payable_by(v["pay_date"], limit):
                db_mark_paid(uid, key, norm_date(v["pay_date"]))
                marked.append(norm_date(v["pay_date"]))
    if not marked:
        await update.message.reply_text(
            "ℹ️ " + r["name"] + "\n"
            "Немає виплат що настали сьогодні або раніше.\n"
            "Для дострокової виплати вкажи дату:\n"
            "/paid " + r.get("account","") + " " + r["veds"][0]["pay_date"] if r["veds"] else ""
        )
        return
    dates_str = ", ".join(marked)
    await update.message.reply_text("✅ Виплачено: " + r["name"] + "\n   🔢 " + r.get("account","") + "\n   📅 " + dates_str)

async def cmd_unpaid_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає.")
        return
    if not ctx.args:
        await update.message.reply_text("⚠️ Вкажи номер рахунку:\n/unpaid 001256811781\n/unpaid 001256811781 05.06")
        return
    query = ctx.args[0].strip()
    specific_date = norm_date(ctx.args[1]) if len(ctx.args) >= 2 else None
    all_rows = build_all_rows(veds, phones, uid=uid)
    found = find_by_account_or_name(all_rows, query)
    if not found:
        await update.message.reply_text("🔍 Не знайдено: " + query)
        return
    if len(found) > 1:
        names = "\n".join([str(i+1) + ". " + r.get("account","") + " " + r["name"] for i, r in enumerate(found)])
        await update.message.reply_text("Знайдено кілька — вкажи номер рахунку:\n" + names)
        return
    r = found[0]
    key = make_person_key(r)
    db_unmark_paid(uid, key, specific_date)
    if specific_date:
        await update.message.reply_text("↩️ Знято позначку: " + r["name"] + " · " + specific_date)
    else:
        await update.message.reply_text("↩️ Знято всі позначки: " + r["name"])

async def cmd_unpaid_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    phones = db_get_phones(uid)
    if not veds:
        await update.message.reply_text("📭 Відомостей немає.")
        return
    all_rows = build_all_rows(veds, phones, uid=uid)
    paid_set = db_get_paid(uid)
    unpaid = []
    for r in all_rows:
        unpaid_veds = [v for v in r["veds"] if not is_ved_paid(r, v, paid_set)]
        if unpaid_veds:
            unpaid.append(dict(r, veds=unpaid_veds))
    unpaid.sort(key=lambda r: r["veds"][0].get("pay_date","99.99"))
    if not unpaid:
        await update.message.reply_text("🎉 Всі виплати здійснено!")
        return
    total = sum(sum(float(v.get("sum",0)) for v in r["veds"]) for r in unpaid)
    await update.message.reply_text("⏳ Не виплачено · " + str(len(unpaid)) + " осіб · " + fmt_hrn(total) + "\n" + "─"*28)
    for chunk in format_rows(unpaid, phones, paid_set, uid=uid):
        await update.message.reply_text(chunk)

async def cmd_clear_paid(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    db_clear_paid(uid)
    await update.message.reply_text("↩️ Всі позначки виплат знято")

async def cmd_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    veds = db_get_vedomosti(uid)
    db_clear_vedomosti(uid)
    await update.message.reply_text("🗑 Очищено " + str(len(veds)) + " відомостей. Телефонна книга збережена.")

async def cmd_phones(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    phones = db_get_phones(uid)
    if phones:
        await update.message.reply_text("📞 Телефонна книга: " + str(len(phones)) + " адрес")
    else:
        await update.message.reply_text("📞 Телефонна книга не завантажена\nНадішли XLSX файл.")

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    uid = update.effective_user.id
    is_xlsx = (doc.mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    ) or doc.file_name.lower().endswith(('.xlsx','.xls')))

    if is_xlsx:
        msg = await update.message.reply_text("⏳ Завантажую телефонну книгу...")
        try:
            file = await ctx.bot.get_file(doc.file_id)
            data = await file.download_as_bytearray()
            phone_map = parse_xlsx(bytes(data))
            db_save_phones(uid, phone_map)
            try:
                n = pb_import_xlsx(uid, bytes(data))
                await msg.edit_text("✅ Телефонна книга: " + str(len(phone_map)) + " адрес · " + str(n) + " записів збережено")
            except Exception as pb_err:
                await msg.edit_text("✅ Телефонна книга: " + str(len(phone_map)) + " адрес збережено")
        except Exception as e:
            await msg.edit_text("❌ Помилка XLSX: " + str(e)[:300])
        return

    if doc.mime_type != "application/pdf":
        await update.message.reply_text("⚠️ Надішли PDF відомість або XLSX телефонну книгу")
        return

    msg = await update.message.reply_text("⏳ Завантажую " + doc.file_name + "...")
    try:
        file = await ctx.bot.get_file(doc.file_id)
        pdf_bytes = await file.download_as_bytearray()
        b64 = base64.b64encode(pdf_bytes).decode()
        await msg.edit_text("🤖 Обробляю через Claude (" + str(len(pdf_bytes)//1024) + " KB)...")
        result = await ask_claude(b64)
        result["source_file"] = doc.file_name
        rows = result.get("rows", [])
        if not rows:
            await msg.edit_text("⚠️ Рядків не знайдено")
            return
        db_save_vedomist(uid, result)
        veds = db_get_vedomosti(uid)
        phones = db_get_phones(uid)
        total_people = len(build_all_rows(veds, phones, uid=uid))
        ved = result.get("vedomist_type","?")
        dil = result.get("dilinitsa","?")
        per = result.get("period","?")
        await msg.edit_text(
            "✅ Збережено: В" + ved + "·Діл." + dil + " · " + per + " · " + str(len(rows)) + " осіб\n\n"
            "📊 Всього відомостей: " + str(len(veds)) + "\n"
            "👥 Унікальних осіб: " + str(total_people) + "\n\n"
            "Команди: /list · /today 07.06 · /status"
        )
    except Exception as e:
        await msg.edit_text("❌ Помилка: " + str(e)[:500])

async def handle_other(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text.strip()
    if text == "/":
        return
    text = text.upper()
    if len(text) >= 2:
        veds = db_get_vedomosti(uid)
        phones = db_get_phones(uid)
        if veds:
            found = [r for r in build_all_rows(veds, phones, uid=uid) if text in r["name"].upper()]
            if found:
                paid_set = db_get_paid(uid)
                await update.message.reply_text("🔍 '" + update.message.text.strip() + "' · " + str(len(found)) + " осіб\n" + "─"*28)
                for chunk in format_rows(found, phones, paid_set, uid=uid):
                    await update.message.reply_text(chunk)
                return
    await update.message.reply_text(
        "📄 Надішли PDF або XLSX\n\n"
        "Команди: /list · /today 07.06 · /status · /clear\n"
        "/search Шв або просто напиши прізвище"
    )


# ════════════════════════════════════════════════════════════════
# PHONE BOOK
# ════════════════════════════════════════════════════════════════

def pb_init(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS pb_entries (
                id SERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL,
                street TEXT NOT NULL,
                name TEXT,
                phone TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            );
            CREATE INDEX IF NOT EXISTS pb_street_idx ON pb_entries(user_id, street);
        """)
    conn.commit()

def pb_import_xlsx(uid, data: bytes):
    import io as _io
    import openpyxl as _xl
    wb = _xl.load_workbook(_io.BytesIO(data))
    ws = wb.active
    entries = []
    current_street = None
    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip()
        col2 = str(row[1] or '').strip()
        col3 = str(row[2] or '').strip()
        col4 = str(row[3] or '').strip()
        col5 = str(row[4] or '').strip()
        if col1 and not col2 and not col3:
            continue
        if col2.lower().startswith('вул.'):
            current_street = col2
        if current_street and (col3 or col4):
            entries.append((uid, current_street, col3 or None, col4 or None, col5 or None))
    with get_conn() as conn:
        pb_init(conn)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pb_entries WHERE user_id=%s", (uid,))
            if entries:
                cur.executemany(
                    "INSERT INTO pb_entries (user_id, street, name, phone, notes) VALUES (%s,%s,%s,%s,%s)",
                    entries
                )
        conn.commit()
    return len(entries)

def strip_addr(s):
    return re.sub(r'[^а-яіїєґa-z0-9]', '', str(s).lower())

def pb_find_by_street(uid, street_query):
    q_stripped = strip_addr(street_query)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, street, name, phone, notes FROM pb_entries WHERE user_id=%s ORDER BY street, id",
                (uid,)
            )
            all_rows = cur.fetchall()
    return [r for r in all_rows if q_stripped in strip_addr(r[1])]

def pb_find_by_name(uid, name_query):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, street, name, phone, notes FROM pb_entries WHERE user_id=%s AND LOWER(name) LIKE %s ORDER BY street, id",
                (uid, '%' + name_query.lower() + '%')
            )
            return cur.fetchall()

def pb_add_entry(uid, street, name, phone, notes=None):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO pb_entries (user_id, street, name, phone, notes) VALUES (%s,%s,%s,%s,%s) RETURNING id",
                (uid, street, name, phone, notes)
            )
            new_id = cur.fetchone()[0]
        conn.commit()
    return new_id

def pb_update_entry(uid, entry_id, name=None, phone=None, notes=None):
    with get_conn() as conn:
        with conn.cursor() as cur:
            if name is not None:
                cur.execute("UPDATE pb_entries SET name=%s, updated_at=NOW() WHERE id=%s AND user_id=%s", (name, entry_id, uid))
            if phone is not None:
                cur.execute("UPDATE pb_entries SET phone=%s, updated_at=NOW() WHERE id=%s AND user_id=%s", (phone, entry_id, uid))
            if notes is not None:
                cur.execute("UPDATE pb_entries SET notes=%s, updated_at=NOW() WHERE id=%s AND user_id=%s", (notes, entry_id, uid))
        conn.commit()

def pb_delete_entry(uid, entry_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM pb_entries WHERE id=%s AND user_id=%s", (entry_id, uid))
        conn.commit()

def pb_export_xlsx(uid) -> bytes:
    import io as _io
    import openpyxl as _xl
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT street, name, phone, notes FROM pb_entries WHERE user_id=%s ORDER BY street, id",
                (uid,)
            )
            rows = cur.fetchall()
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Телефонна книга"
    ws.append(['п/п', 'Вулиця/№будинку', 'ПІБ', '№ телефону', 'Примітки'])
    current_street = None
    for street, name, phone, notes in rows:
        if street != current_street:
            current_street = street
            ws.append([None, street, name, phone, notes])
        else:
            ws.append([None, None, name, phone, notes])
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def fmt_pb_entries(entries):
    if not entries:
        return "Нічого не знайдено"
    lines = []
    current_street = None
    for eid, street, name, phone, notes in entries:
        if street != current_street:
            current_street = street
            lines.append("\n📍 " + street)
        line = "  [" + str(eid) + "] " + (name or '—')
        if phone: line += " · 📞 " + phone
        if notes: line += " · 📝 " + notes
        lines.append(line)
    return "\n".join(lines).strip()

# ── Phone book handlers ────────────────────────────────────────────────────────

async def cmd_pb_find(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ctx.args:
        await update.message.reply_text(
            "🔍 Пошук у телефонній книзі:\n"
            "/pb_find Шкільна 7 — по адресі\n"
            "/pb_find Левчук — по ПІБ"
        )
        return
    query = " ".join(ctx.args).strip()
    if any(c.isdigit() for c in query) or 'вул' in query.lower():
        entries = pb_find_by_street(uid, query)
    else:
        entries = pb_find_by_name(uid, query)
        if not entries:
            entries = pb_find_by_street(uid, query)
    if not entries:
        await update.message.reply_text("🔍 Нічого не знайдено: " + query + "\n\nДодати: /pb_phone " + query + " 0991234567")
        return
    text = "🔍 Знайдено " + str(len(entries)) + " записів:\n" + fmt_pb_entries(entries)
    text += "\n\nРедагувати: /pb_edit [ID] [поле] [значення]\nВидалити: /pb_del [ID]"
    await update.message.reply_text(text[:4000])

async def cmd_pb_add(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ctx.args:
        await update.message.reply_text(
            "➕ Додати запис:\n"
            "/pb_add вул. Шкільна, 7 | Іван Петрович | 0991234567\n"
            "/pb_add вул. Шкільна, 7 | Марія | 0661234567 | примітка"
        )
        return
    text = " ".join(ctx.args)
    parts = [p.strip() for p in text.split("|")]
    if len(parts) < 3:
        await update.message.reply_text("⚠️ Формат: /pb_add вул. Адреса | ПІБ | телефон")
        return
    street = parts[0]
    name   = parts[1]
    phone  = parts[2]
    notes  = parts[3] if len(parts) > 3 else None
    new_id = pb_add_entry(uid, street, name, phone, notes)
    await update.message.reply_text(
        "✅ Додано [" + str(new_id) + "]:\n"
        "📍 " + street + "\n"
        "👤 " + name + "\n"
        "📞 " + phone +
        ("\n📝 " + notes if notes else "")
    )

async def cmd_pb_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/pb_phone Зарічна 36 0991234567 — швидко додати телефон по адресі"""
    uid = update.effective_user.id
    if not ctx.args or len(ctx.args) < 2:
        await update.message.reply_text(
            "📞 Швидке додавання телефону:\n"
            "/pb_phone Зарічна 36 0991234567\n"
            "/pb_phone Ковельська 174 0671234567\n\n"
            "Останній аргумент — телефон, решта — адреса"
        )
        return

    phone = ctx.args[-1]
    addr_parts = ctx.args[:-1]

    # Перевірити що останній аргумент схожий на телефон
    if not re.match(r'^[\d\+\(\)\-]{7,}$', phone):
        await update.message.reply_text(
            "⚠️ Останній аргумент має бути телефоном\n"
            "Приклад: /pb_phone Зарічна 36 0991234567"
        )
        return

    addr_query = " ".join(addr_parts)

    # Спробувати знайти існуючий запис за адресою
    existing = pb_find_by_street(uid, addr_query)

    if existing:
        # Знайдено — показати і запропонувати оновити перший без телефону
        no_phone = [e for e in existing if not e[3]]
        if no_phone:
            # Оновити перший запис без телефону
            entry = no_phone[0]
            pb_update_entry(uid, entry[0], phone=phone)
            await update.message.reply_text(
                "✅ Телефон додано:\n"
                "📍 " + entry[1] + "\n"
                "👤 " + (entry[2] or '—') + "\n"
                "📞 " + phone + "\n"
                "[ID: " + str(entry[0]) + "]"
            )
        else:
            # Всі записи вже мають телефон — додати новий
            # Визначити вулицю з першого знайденого запису
            street = existing[0][1]
            new_id = pb_add_entry(uid, street, None, phone)
            await update.message.reply_text(
                "➕ Додано новий запис:\n"
                "📍 " + street + "\n"
                "📞 " + phone + "\n"
                "[ID: " + str(new_id) + "]\n\n"
                "ℹ️ Існуючі записи за цією адресою вже мали телефони.\n"
                "Редагувати ПІБ: /pb_edit " + str(new_id) + " name Прізвище Ім'я"
            )
    else:
        # Не знайдено — створити новий запис
        # Сформувати вулицю з адреси
        street = "вул. " + addr_query
        new_id = pb_add_entry(uid, street, None, phone)
        await update.message.reply_text(
            "➕ Новий запис додано:\n"
            "📍 " + street + "\n"
            "📞 " + phone + "\n"
            "[ID: " + str(new_id) + "]\n\n"
            "Додати ПІБ: /pb_edit " + str(new_id) + " name Прізвище Ім'я\n"
            "Перевірити адресу: /pb_find " + addr_query
        )

async def cmd_pb_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ctx.args or len(ctx.args) < 3:
        await update.message.reply_text(
            "✏️ Редагувати запис:\n"
            "/pb_edit [ID] phone 0991234567\n"
            "/pb_edit [ID] name Нове ПІБ\n"
            "/pb_edit [ID] notes примітка\n\n"
            "ID видно після /pb_find"
        )
        return
    try:
        entry_id = int(ctx.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ ID має бути числом")
        return
    field = ctx.args[1].lower()
    value = " ".join(ctx.args[2:])
    if field in ('phone', 'телефон'):
        pb_update_entry(uid, entry_id, phone=value)
        await update.message.reply_text("✅ Телефон оновлено: [" + str(entry_id) + "] → " + value)
    elif field in ('name', 'піб', 'пib'):
        pb_update_entry(uid, entry_id, name=value)
        await update.message.reply_text("✅ ПІБ оновлено: [" + str(entry_id) + "] → " + value)
    elif field in ('notes', 'примітка'):
        pb_update_entry(uid, entry_id, notes=value)
        await update.message.reply_text("✅ Примітку оновлено: [" + str(entry_id) + "] → " + value)
    else:
        await update.message.reply_text("⚠️ Поле: phone, name або notes")

async def cmd_pb_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ctx.args:
        await update.message.reply_text("🗑 Видалити запис: /pb_del [ID]")
        return
    try:
        entry_id = int(ctx.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ ID має бути числом")
        return
    pb_delete_entry(uid, entry_id)
    await update.message.reply_text("🗑 Запис [" + str(entry_id) + "] видалено")

async def cmd_pb_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    msg = await update.message.reply_text("⏳ Формую файл...")
    try:
        data = pb_export_xlsx(uid)
        await update.message.reply_document(
            document=data,
            filename="телефонна_книга.xlsx",
            caption="📞 Актуальна телефонна книга"
        )
        await msg.delete()
    except Exception as e:
        await msg.edit_text("❌ Помилка: " + str(e)[:300])

async def cmd_link(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/link 907540199004 748 — прив'язати рахунок до запису телефонної книги"""
    uid = update.effective_user.id
    if not ctx.args or len(ctx.args) < 2:
        await update.message.reply_text(
            "🔗 Прив'язати рахунок до телефонної книги:\n"
            "/link [рахунок] [ID запису]\n\n"
            "Приклад:\n"
            "/link 907540199004 748\n\n"
            "ID запису видно після /pb_find"
        )
        return
    account = ctx.args[0].strip()
    try:
        entry_id = int(ctx.args[1])
    except ValueError:
        await update.message.reply_text("⚠️ ID запису має бути числом")
        return

    # Перевірити що запис існує
    entry = db_get_pb_entry(uid, entry_id)
    if not entry:
        await update.message.reply_text("⚠️ Запис [" + str(entry_id) + "] не знайдено в телефонній книзі")
        return

    db_link_account(uid, account, entry_id)
    phone = entry[3] or "—"
    name  = entry[2] or "—"
    await update.message.reply_text(
        "🔗 Прив'язано:\n"
        "🔢 Рахунок: " + account + "\n"
        "📍 " + entry[1] + "\n"
        "👤 " + name + "\n"
        "📞 " + phone + "\n"
        "[ID: " + str(entry_id) + "]"
    )

async def cmd_unlink(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/unlink 907540199004 — зняти прив'язку рахунку"""
    uid = update.effective_user.id
    if not ctx.args:
        await update.message.reply_text("🔗 Зняти прив'язку: /unlink [рахунок]")
        return
    account = ctx.args[0].strip()
    db_unlink_account(uid, account)
    await update.message.reply_text("↩️ Прив'язку знято для рахунку: " + account)

async def cmd_links(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/links — показати всі прив'язки"""
    uid = update.effective_user.id
    links = db_get_links(uid)
    if not links:
        await update.message.reply_text(
            "🔗 Прив'язок немає.\n\n"
            "Додати: /link [рахунок] [ID з /pb_find]"
        )
        return
    lines = ["🔗 Прив'язки рахунків (" + str(len(links)) + "):"]
    for account, entry_id in sorted(links.items()):
        entry = db_get_pb_entry(uid, entry_id)
        if entry:
            phone = entry[3] or "—"
            name  = entry[2] or "—"
            lines.append("  " + account + " → [" + str(entry_id) + "] " + name + " · 📞 " + phone)
        else:
            lines.append("  " + account + " → [" + str(entry_id) + "] (запис видалено)")
    await update.message.reply_text("\n".join(lines)[:4000])

def main():
    init_db()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start",       start))
    app.add_handler(CommandHandler("status",      cmd_status))
    app.add_handler(CommandHandler("list",        cmd_list))
    app.add_handler(CommandHandler("today",       cmd_today))
    app.add_handler(CommandHandler("search",      cmd_search))
    app.add_handler(CommandHandler("multi",       cmd_multi))
    app.add_handler(CommandHandler("clear",       cmd_clear))
    app.add_handler(CommandHandler("phones",      cmd_phones))
    app.add_handler(CommandHandler("paid",        cmd_paid))
    app.add_handler(CommandHandler("unpaid",      cmd_unpaid_cmd))
    app.add_handler(CommandHandler("unpaid_list", cmd_unpaid_list))
    app.add_handler(CommandHandler("clear_paid",  cmd_clear_paid))
    app.add_handler(CommandHandler("link",        cmd_link))
    app.add_handler(CommandHandler("unlink",      cmd_unlink))
    app.add_handler(CommandHandler("links",       cmd_links))
    app.add_handler(CommandHandler("pb_find",     cmd_pb_find))
    app.add_handler(CommandHandler("pb_add",      cmd_pb_add))
    app.add_handler(CommandHandler("pb_phone",    cmd_pb_phone))
    app.add_handler(CommandHandler("pb_edit",     cmd_pb_edit))
    app.add_handler(CommandHandler("pb_del",      cmd_pb_del))
    app.add_handler(CommandHandler("pb_export",   cmd_pb_export))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_other))
    print("Bot started with DB")
    app.run_polling(close_loop=False)

if __name__ == "__main__":
    main()
